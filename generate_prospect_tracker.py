import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: Oil Trading Prospects
# ============================================================
ws1 = wb.active
ws1.title = "Oil Trading Prospects"

# Styles
header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
tier1_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
tier2_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
tier3_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["#", "Tier", "Company", "MD / Key Contact", "Phone", "Email", "Website",
           "Type", "Products to Pitch", "Status", "Date Contacted", "Response", "Follow-up Date", "Notes"]

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Column widths
widths = [4, 6, 22, 25, 18, 30, 22, 16, 20, 12, 14, 14, 14, 20]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Data
prospects = [
    # Tier 1
    [1, "T1", "Arc Energy Trading", "Jing Goh / Ben Tan (Trading Mgrs)", "+65 6836 1018", "Contact form", "arcenergy.sg", "Cargo & bunker", "Mini CTRM + LC", "", "", "", "", "Small, est. 2014, NUS founder"],
    [2, "T1", "TIS Petroleum Asia", "Ivan Handojo (Founder/MD)", "+65 6291 3507", "Contact form", "tispetrol.com", "Crude & upstream", "All 3", "", "", "", "", "Founder-led, small"],
    [3, "T1", "Blue Ocean Int'l Oil", "Oliver Wolter (MD)", "+65 8228 2614", "blueoceanmineraloel.com", "blueoceanmineraloel.com", "Oil products", "Mini CTRM + E2O", "", "", "", "", "MD mobile available"],
    [4, "T1", "Capricorn Commodities", "Capt. Sachin Saxena (Founder/MD)", "+65 9621 0539 (WA)", "saxenaskp@capricorn-singapore.com", "capricorn-singapore.com", "Commodity/shipping", "All 3", "", "", "", "", "SME 500 winner, WhatsApp available"],
    [5, "T1", "Khong Lieng Trading", "Edward Quek (Group COO)", "+65 9384 3844", "edward@khonglieng.com.sg", "khonglieng.com.sg", "Marine, O&G, refinery", "LC + E2O", "", "", "", "", "Est. 1972, 50+ years"],
    [6, "T1", "Tropical Oil ACI", "Daniel Lim (MD)", "via LinkedIn", "LinkedIn DM", "tropicaloil.com.sg", "Oil trading", "Mini CTRM", "", "", "", "", "LinkedIn outreach only"],
    # Tier 2
    [7, "T2", "Montfort Group", "Ex-Shell/Trafigura founders", "+65 3105 1583", "singapore@mont-fort.com", "mont-fort.com", "Physical & downstream", "Mini CTRM + LC", "", "", "", "", "Growing, ex-majors team"],
    [8, "T2", "PetroEast Singapore", "Jianwei Zhang (MD)", "+65 6361 2200", "Contact form", "petroeast.com.sg", "Crude & products", "Mini CTRM + E2O", "", "", "", "", "Est. 2003"],
    [9, "T2", "Clipper Oil", "Zheng Tian (Founder)", "+65 3129 1232", "bunkers@clipperoil.com", "clipperoil.com", "Marine fuel", "E2O + LC", "", "", "", "", "Founder accessible"],
    [10, "T2", "Bunker House Petroleum", "Colin Wang", "+65 6222 0337", "bunkerhse.com.sg", "bunkerhse.com.sg", "Bunker (Top 10)", "E2O + LC", "", "", "", "", "Top 10 MPA supplier"],
    [11, "T2", "Straits Bunkering", "Shafiq Nezammuddin (Director)", "via website", "operations@straitsbunkering.com", "straitsbunkering.com", "Bunker", "E2O + Mini CTRM", "", "", "", "", "Director on LinkedIn"],
    [12, "T2", "Consort Bunkers", "General", "+65 6344 3008", "bunkers@consortbunkers.com.sg", "consortbunkers.com.sg", "Multi-port bunker", "E2O + LC", "", "", "", "", "Multi-port, growing"],
    [13, "T2", "Seven Seas Oil Trading", "General", "+65 6227 3338", "supply.singapore@sevenseasgroup.com", "sevenseasgroup.com", "Bunker, 30+ years", "Mini CTRM + LC", "", "", "", "", "Est. 1992"],
    [14, "T2", "Island Oil Trading", "Team on website", "+65 6653 1941", "ops@island-oil.com", "island-oil.com", "Bunker, global", "E2O + LC", "", "", "", "", "Global operations"],
    [15, "T2", "Cathay Petroleum", "General", "via website", "Form at cathaypetroleum.com", "cathaypetroleum.com", "Crude & products", "Mini CTRM", "", "", "", "", "HK HQ, SG office"],
    # Tier 3
    [16, "T3", "Sirius Marine", "General", "+65 6444 2900", "operations@siriusoil.com.sg", "siriusgroupinternational.com", "Marine fuel logistics", "E2O", "", "", "", "", "Est. 1996"],
    [17, "T3", "ENH Energy Trading", "Vitol JV", "via website", "enhenergytrading.com", "enhenergytrading.com", "LNG, LPG, condensate", "Mini CTRM", "", "", "", "", "Vitol-backed"],
    [18, "T3", "Triton Bunkering", "General", "+65 6440 2522", "triton@tritonbunkering.com", "tritonbunkering.com", "Bunker", "E2O + LC", "", "", "", "", ""],
    [19, "T3", "Palmstone Tankers", "General", "+65 6473 2460", "via website", "palmstone", "Tankers & trading", "E2O", "", "", "", "", ""],
    [20, "T3", "Vegatron", "General", "+65 6777 7772", "info@vegatron.com.sg", "vegatron.com.sg", "Diesel & lubricants", "E2O", "", "", "", "", ""],
    [21, "T3", "Enigma Energy", "General", "via website", "enigmaenergies.com", "enigmaenergies.com", "Fuel oil & petroleum", "Mini CTRM", "", "", "", "", ""],
    [22, "T3", "Ocean Energy", "General", "via website", "oceanenergy.com.sg", "oceanenergy.com.sg", "Oil products", "Mini CTRM", "", "", "", "", ""],
    [23, "T3", "Sea Splendor Int'l", "General", "via website", "sg_bunkers@seatrader.hk", "seasplendorpetro.com", "Bunker", "E2O + LC", "", "", "", "", ""],
    [24, "T3", "Commodity Connect", "Capt. Sohail Mahbub (MD)", "via LinkedIn", "LinkedIn DM", "via LinkedIn", "Commodity trading", "Mini CTRM", "", "", "", "", ""],
    [25, "T3", "Pegasus Maritime", "General", "+65 9692 5999", "pegasus-m.com", "pegasus-m.com", "Bunker", "E2O", "", "", "", "", "MPA licensed"],
    [26, "T3", "Dan-Bunkering SG", "General", "via website", "dan-bunkering.com", "dan-bunkering.com", "Bunker", "E2O", "", "", "", "", "Danish company"],
    [27, "T3", "Sentek Marine", "General", "+65 6692 0528", "MPA list", "sentek.com.sg", "Bunker", "E2O + LC", "", "", "", "", ""],
    [28, "T3", "Integr8 Fuels SG", "General", "via website", "integr8fuels.com", "integr8fuels.com", "Marine fuel", "E2O", "", "", "", "", ""],
    [29, "T3", "Pacific Petroleum", "General", "via website", "pacific-petroleum.asia", "pacific-petroleum.asia", "Petroleum", "Mini CTRM", "", "", "", "", ""],
    [30, "T3", "Costank (S)", "General", "MPA list", "MPA list", "MPA list", "Bunker", "E2O", "", "", "", "", ""],
    [31, "T3", "Searights Maritime", "General", "MPA list", "MPA list", "MPA list", "Bunker", "E2O", "", "", "", "", ""],
    [32, "T3", "Alliance Oil Trading", "General", "MPA list", "MPA list", "MPA list", "Oil trading", "Mini CTRM", "", "", "", "", ""],
    [33, "T3", "Global Energy Trading", "General", "MPA list", "MPA list", "MPA list", "Bunker", "E2O", "", "", "", "", ""],
    [34, "T3", "Universal Energy", "General", "MPA list", "MPA list", "MPA list", "Bunker", "E2O", "", "", "", "", ""],
    [35, "T3", "PetroChina Int'l (S)", "General", "+65 6411 7531", "MPA list", "petrochina.com", "Petroleum", "Mini CTRM", "", "", "", "", "State-owned subsidiary"],
]

for row_idx, data in enumerate(prospects, 2):
    tier = data[1]
    fill = tier1_fill if tier == "T1" else tier2_fill if tier == "T2" else tier3_fill
    for col_idx, val in enumerate(data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = fill
        cell.alignment = wrap
        cell.border = thin_border

# ============================================================
# SHEET 2: Accounting Firm Partners
# ============================================================
ws2 = wb.create_sheet("Accounting Firm Partners")

headers2 = ["#", "Firm", "Founder / Contact", "Phone", "Email", "Website",
            "SME Client Base", "Status", "Date Contacted", "Response", "Follow-up Date", "Notes"]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

widths2 = [4, 22, 25, 18, 30, 22, 16, 12, 14, 14, 14, 20]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

acct_firms = [
    [1, "Ackenting Group (AG)", "John Woo (Founder/MD)", "+65 9383 2464", "johnwoo@ag-singapore.com", "ag-singapore.com", "260+ reviews", "", "", "", "", "Tech-forward, open to partnerships"],
    [2, "3E Accounting", "Lawrence Chai (Founder/CTO)", "+65 6690 9266", "info@3ecpa.com.sg", "3ecpa.com.sg", "Regional, 10+ years", "", "", "", "", "Digital-first, writes about tech"],
    [3, "K Cloud Accounting", "Tommy (Founder)", "+65 9863 8665", "tommyksh@kca.sg", "kca.sg", "Startups & SMEs", "", "", "", "", "Cloud-first, small firm"],
    [4, "J Accounting", "Rebekah Tan (Founder)", "via website", "sales@j-accountingservices.com", "j-accountingservices.com", "SMEs", "", "", "", "", "Chartered accountant"],
    [5, "SBS Consulting", "Management team", "+65 6536 0036", "info@sbsgroup.com.sg", "sbsgroup.com.sg", "Large SME base", "", "", "", "", "One-stop corporate services"],
    [6, "Counto", "Saba Khan (Co-founder/COO)", "via website", "Partner page: counto.sg/partners", "counto.sg", "SMEs", "", "", "", "", "Has partner programme"],
    [7, "Accountancy Hub", "Via LinkedIn", "via website", "accountancyhub.com.sg", "accountancyhub.com.sg", "SMEs", "", "", "", "", "Tax, audit, compliance"],
    [8, "ContactOne Consulting", "General", "via website", "contactone.com.sg", "contactone.com.sg", "SMEs", "", "", "", "", "Corp services + accounting"],
]

for row_idx, data in enumerate(acct_firms, 2):
    for col_idx, val in enumerate(data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
        cell.alignment = wrap
        cell.border = thin_border

# ============================================================
# SHEET 3: Email Templates
# ============================================================
ws3 = wb.create_sheet("Outreach Templates")

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 100

templates = [
    ["TEMPLATE", "CONTENT"],
    ["", ""],
    ["TIER 1 EMAIL", ""],
    ["Subject", "Mini CTRM + LC checker — built by ex-Citi commodities tech head"],
    ["Body", """Hi [Name],

I'm Gopalan — most recently I led Trading & Shipping technology at Ampol (covering international oil trading, shipping, finance and treasury systems), and before that spent 8 years as Asia Head of Commodities Trading & Risk Technology at Citi, covering oil, gas, ags, and exotics across the region.

I've taken that experience and built three tools specifically for trading houses that don't need a $500K Openlink/Allegro setup but have outgrown spreadsheets:

1. Mini CTRM for Oil
Lightweight trade capture, position tracking, and P&L — designed for physical oil desks running 50-500 trades/month. Up and running in weeks, not months.

2. LC Compliance Checker
Upload your LC docs (PDF or scanned). AI extracts every field, cross-checks against UCP600 rules, and flags discrepancies by severity — critical, major, minor — with confidence scores. What takes your ops team 2-3 hours now takes 5 minutes.

3. Email-to-Order Dashboard
AI reads incoming order/nomination emails, extracts deal terms (product, quantity, pricing, dates, counterparty), and populates a live dashboard. No more copy-pasting from emails into spreadsheets.

I know how oil trading desks actually work — the 2am cargo nominations, the LC discrepancy panic, the spreadsheet that "only one person knows how to use." These tools solve real problems I've seen firsthand.

Would 15 minutes for a demo be worth your time? No sales pitch — I'll show you the product running on real data.

Best regards,
Gopalan L S
+65 9180 2686 | jskai.ai
Ex-Ampol T&S Tech | Ex-Citi Asia Head, Commodities Trading & Risk Technology"""],
    ["", ""],
    ["TIER 2 EMAIL", ""],
    ["Subject", "Automating LC checks and email orders for oil trading desks"],
    ["Body", """Hi,

I'm reaching out because I've built automation tools specifically for oil trading operations — and I think they'd save your team significant time.

Quick background: I most recently led Trading & Shipping technology at Ampol (international oil trading operations), and before that spent 8 years as Asia Head of Commodities Trading & Risk Technology at Citigroup. I know the operational pain of oil trading desks firsthand.

Three tools I've built:

LC Compliance Checker — Upload LC documents (PDF/scan), AI checks every field against UCP600 rules and flags discrepancies. Cuts review from hours to minutes.

Email-to-Order Dashboard — AI reads order/nomination emails and extracts deal terms into a live dashboard. Eliminates manual data entry.

Mini CTRM — Lightweight trade capture and position management for physical oil desks. Fraction of the cost of enterprise solutions.

Happy to do a 15-minute demo for your trading or operations team. Could you point me to the right person?

Best regards,
Gopalan L S
+65 9180 2686 | jskai.ai"""],
    ["", ""],
    ["LINKEDIN MESSAGE", ""],
    ["For MDs", "Hi [Name], I led T&S tech at Ampol and was Asia Head of Commodities Trading Tech at Citi for 8 years. I've built a lightweight CTRM, LC compliance checker, and email-to-order tool for small/mid oil trading houses. Built these because I've lived the pain. Would a 15-min demo be worth your time?"],
    ["", ""],
    ["WHATSAPP MESSAGE", ""],
    ["For mobile contacts", """Hi [Name], I'm Gopalan — I led T&S tech at Ampol and was Asia Head of Commodities Trading Tech at Citi. I've built a mini CTRM, LC compliance checker, and email-to-order dashboard specifically for oil trading desks. Would love to show you a quick 15-min demo. No commitment — just the product on real data. Let me know if you're open to it. jskai.ai"""],
    ["", ""],
    ["ACCOUNTING FIRM EMAIL", ""],
    ["Subject", "Partnership idea — AI automation for your SME clients"],
    ["Body", """Hi [Name],

I run JSK AI (jskai.ai) — we help SMEs automate workflows like invoicing, customer enquiries, and document processing using AI.

I'm reaching out because your clients are exactly the businesses we serve, and I think there's a simple partnership that benefits both of us:

For your clients: We automate their time-consuming processes (invoice handling, data entry, customer follow-ups) — saving them 20+ hours/week. This makes them more organised, which makes your job easier at tax time.

For your firm: We offer a referral fee of SGD 200-500 per client, or a revenue share on ongoing automation subscriptions. Zero effort on your end — you just make the introduction.

For context: I spent 20 years in financial services tech (Citi, DBS) and now apply that to practical AI automation for small businesses. No jargon, no long contracts — just results.

Would a 15-minute call this week make sense? Happy to walk you through what we've built.

Gopalan L S
JSK AI Automation & Digital Solutions
+65 9180 2686 | jskai.ai"""],
]

for row_idx, data in enumerate(templates, 1):
    for col_idx, val in enumerate(data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = wrap
        cell.border = thin_border
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
        elif data[0] and data[0].isupper() and "EMAIL" in data[0] or "MESSAGE" in data[0] or "TEMPLATE" in data[0]:
            cell.font = Font(bold=True, size=12, color="6366F1")
        elif data[0] == "Subject":
            cell.font = Font(bold=True)

# Freeze panes
ws1.freeze_panes = "A2"
ws2.freeze_panes = "A2"
ws3.freeze_panes = "A2"

# Auto-filter
ws1.auto_filter.ref = f"A1:N{len(prospects)+1}"
ws2.auto_filter.ref = f"A1:L{len(acct_firms)+1}"

output_path = os.path.join(r"d:\Claude Projects\JSK AIAutomation and Digital Solutions", "JSK_AI_Prospect_Tracker.xlsx")
wb.save(output_path)
print(f"Saved to: {output_path}")
